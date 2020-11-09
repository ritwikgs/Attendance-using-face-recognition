from tkinter import *
import os
from PIL import Image, ImageTk
import openpyxl
from openpyxl import load_workbook
from datetime import date
from twilio.rest import Client


class GUI(Tk):
    def __init__(self):
        super(GUI, self).__init__()
        self.geometry("800x500")
        self.config(bg="silver")
        self.resizable(0, 0)
        self.title("Attendance system")

    def create_menu(self):
        self.mymenu = Menu()
        self.mymenu.add_cascade(label="Attendance", command=f1)
        self.mymenu.add_cascade(label="Students", command=f1)

        m3 = Menu(self.mymenu, tearoff=0)
        m3.add_command(label="About us")
        self.config(menu=self.mymenu)
        self.mymenu.add_cascade(label="More", menu=m3)


def f1():
    print("ok")


def write_to_txt(name_val, usn_val, sec_val, sem_val, phone_val, email_val):
    print("writting")
    path = 'students'
    if not os.path.exists('students'):
        os.makedirs('students')

    elif not os.path.exists(f"{path}/{sem_val}"):
        os.makedirs(f"{path}/{sem_val}/{sec_val}")
    else:
        file = open(f"{path}/{sem_val}/{sec_val}.txt", "a")
        file.write(f"{usn_val},{name_val},{sem_val},{sec_val},{phone_val},{email_val}\n")
        file.close()


def create_xls(path):
    wb = openpyxl.Workbook()
    if path == "student.xlsx":
        wb.create_sheet("year1")
        wb.create_sheet("year2")
        wb.create_sheet("year3")
        wb.create_sheet("year4")
        wb.save(path)
    else:
        wb.save(f"Attendance\{path}.xlsx")


def write_to_xls(name, usn, sec, sem, phone, email):
    if sem == "1" or sem == "2":
        if sec == "a":
            year = "year1_a"
        else:
            year = "year1_b"

    elif sem == "3" or sem == "4":
        if sec == "a":
            year = "year2_a"
        else:
            year = "year2_b"
    elif sem == "5" or sem == "6":
        if sec == "a":
            year = "year3_a"
        else:
            year = "year3_b"
    else:
        if sec == "a":
            year = "year4_a"
        else:
            year = "year4_b"
    wb = load_workbook(f"student.xlsx")
    sheet = wb.get_sheet_by_name(f"{year}")
    sheet.cell(row=1, column=1).value = "Sl.NO"
    sheet.cell(row=1, column=2).value = "Name"
    sheet.cell(row=1, column=3).value = "Enrollment"
    sheet.cell(row=1, column=4).value = "Section"
    sheet.cell(row=1, column=5).value = "Semester"
    sheet.cell(row=1, column=6).value = "Contact No"
    sheet.cell(row=1, column=7).value = "Email id"
    sheet.cell(row=2, column=1).value = "1"
    c = 0
    current_row = sheet.max_row
    current_column = sheet.max_column

    sheet.cell(row=current_row, column=2).value = name
    sheet.cell(row=current_row, column=3).value = usn
    sheet.cell(row=current_row, column=4).value = sec
    sheet.cell(row=current_row, column=5).value = sem
    sheet.cell(row=current_row, column=6).value = phone
    sheet.cell(row=current_row, column=7).value = email
    sheet.cell(row=current_row + 1, column=1).value = int(sheet.cell(current_row, 1).value) + 1

    wb.save('student.xlsx')
    path = f"Attendance/{year}.xlsx"
    if not os.path.exists(path):
        create_xls(path)
    wb = load_workbook(f"Attendance/{year}.xlsx")
    sheet = wb.active
    sheet.cell(row=1, column=1).value = "Sl.NO"
    sheet.cell(row=1, column=2).value = "Name"
    sheet.cell(row=1, column=3).value = "Enrollment"
    sheet.cell(row=1, column=4).value = " "
    sheet.cell(row=2, column=1).value = "1"
    current_row = sheet.max_row

    sheet.cell(row=current_row, column=2).value = name
    sheet.cell(row=current_row, column=3).value = usn
    sheet.cell(row=current_row + 1, column=1).value = int(sheet.cell(current_row, 1).value) + 1
    wb.save(f'Attendance\{year}.xlsx')


def send_message(name, sub, contact):
    # the following line needs your Twilio Account SID and Auth Token
    client = Client("ACf99b80984b43f1b59121a72685deaba9", "77eb50a8910ec309771c0a9538787a0b")

    # change the "from_" number to your Twilio number and the "to" number
    # to the phone number you signed up for Twilio with, or upgrade your
    # account to send SMS to any phone number
    for i in range(len(name)):
        client.messages.create(to=f"+91{contact[i]}", from_="+1 571 290 2779", body=f"{name[i]} is absent for {sub} class on {date.today()}")
