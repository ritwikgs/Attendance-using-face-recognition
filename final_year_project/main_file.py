from openpyxl import load_workbook
from datetime import date
import exe1
from tkinter import *
import tkinter.messagebox as tmsg
import os
import capture_image as ci
import openpyxl
import train as t
import recognition_video as rv

global a
a = 1
global b
b = 1
global label_list
label_list = []


def change():
    student.pack_forget()
    attendance.pack(anchor="w", side="left", fill="both", expand=TRUE)


def change0():
    attendance.pack_forget()
    student.pack(anchor="w", fill='both', side=LEFT, expand=TRUE)


def change1():
    global a
    if a == 1:
        return
    else:
        view_label.pack_forget()
        attendance_label2.pack(anchor="w", fill="both", side=LEFT, expand=TRUE)
        a = 1


def change2():
    global a
    if a == 0:
        return
    else:
        attendance_label2.pack_forget()
        view_label.pack(anchor="w", fill="both", side=LEFT, expand=TRUE)
        a = 0


def change3():
    global b
    if b == 1:
        return
    else:
        student_label2.pack_forget()
        student_label1.pack(anchor="w", fill="both", side=LEFT, expand=TRUE)
        b = 1
        student_frame1.pack_forget()
        student_frame1.pack(anchor="w", fill=BOTH, side="left")


def change4():
    global b
    if b == 0:
        return
    else:
        student_label1.pack_forget()
        student_label2.pack(anchor="w", fill="both", side=LEFT, expand=TRUE)
        b = 0


def change5():
    fd1.pack_forget()
    fd.pack(pady=60)


def back():
    fd.pack_forget()
    fd1.pack(pady=100)


def write():
    exe1.write_to_txt(name_val.get(), usn_val.get(), sec_val.get(), sem_val.get(), phone_val.get(), email_val.get())
    name.delete(first=0, last=100)
    usn.delete(first=0, last=100)
    section.delete(first=0, last=100)
    semester.delete(first=0, last=100)
    contact.delete(first=0, last=100)
    email.delete(first=0, last=100)


def verify():
    user_pass = l_pass.get()
    user_id = lid.get()
    if user_id == "" and user_pass == "":
        login_by = "Admin"
        login_allow()
    else:
        tmsg.showinfo("", "wrong input")


def login_allow():
    l_login.pack_forget()
    # menu code
    window.mymenu = Menu()
    window.mymenu.add_cascade(label="Attendance", command=change)
    window.mymenu.add_cascade(label="Students", command=change0)
    m3 = Menu(window.mymenu, tearoff=0)
    m3.add_command(label="About us")
    window.config(menu=window.mymenu)
    window.mymenu.add_cascade(label="More", menu=m3)


def write():
    exe1.write_to_txt(name_val.get(), usn_val.get(), sec_val.get(), sem_val.get(), phone_val.get(), email_val.get())
    if not os.path.exists("student.xlsx"):
        path = "student.xlsx"
        exe1.create_xls(path)
    exe1.write_to_xls(name_val.get(), usn_val.get(), sec_val.get(), sem_val.get(), phone_val.get(), email_val.get())
    name.delete(first=0, last=100)
    usn.delete(first=0, last=100)
    section.delete(first=0, last=100)
    semester.delete(first=0, last=100)
    contact.delete(first=0, last=100)
    email.delete(first=0, last=100)


def take_photo():
    ci.detect(usn_val.get(), name_val.get(), sem_val.get())
    tmsg.showinfo(message= "capture successful.\n please train the model.")


def get_student_info():
    # student_frame2.pack_forget()
    student_frame3 = Frame(student_label2, bg="green")
    os.startfile("student.xlsx")
    but2 = Button(student_frame3, text="Back", width="15").pack(side=LEFT, padx="20", pady="10")

    student_frame3.pack(anchor="w", side="left", fill="both", expand=TRUE)


def detect1():
    global label_list
    label_list = rv.detect(a_sem.get(), a_sec.get())


def update1():
    u_sem = a_sem.get()
    u_sec = a_sec.get()
    u_sub = a_sub.get()
    if u_sem == "1" or u_sem == "2":
        if u_sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year1_a.xlsx"
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year1_b.xlsx"

    elif u_sem == "3" or u_sem == "4":
        if u_sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year2_a.xlsx"
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year2_b.xlsx"
    elif u_sem == "5" or u_sem == "6":
        if u_sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year3_a.xlsx"
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year3_b.xlsx"
    else:
        if u_sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year4_a.xlsx"
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year4_b.xlsx"
    global label_list
    if not label_list:
        tmsg.showinfo(message="detect the student first")
        return
    wb = load_workbook(year)
    sheet = wb.get_sheet_by_name(f"{u_sub}")
    col = sheet.max_column
    r = 2
    d = str(date.today())
    if not sheet.cell(row=1, column=col).value == d:
        col = col + 1
        sheet.cell(row=1, column=col).value = d
    while r < sheet.max_row:
        temp = sheet.cell(row=r, column=3).value
        # print(temp)
        usn = temp[3:5]
        usn = usn+temp[-3:]
        # print(temp)
        if int(usn) in label_list:
            sheet.cell(row=r, column=col).value = "present"
        else:
            sheet.cell(row=r, column=col).value = "Absent"
        r = r + 1

    wb.save(year)
    tmsg.showinfo(message="attendance updated")


def show1():
    u_sem = a_sem.get()
    u_sec = a_sec.get()
    if u_sem == "1" or u_sem == "2":
        if u_sec == "a":
            os.startfile(r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance/year1_a.xlsx", )
        else:
            os.startfile(r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance/year1_b.xlsx", )

    elif u_sem == "3" or u_sem == "4":
        if u_sec == "a":
            os.startfile(r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance/year2_a.xlsx", )
        else:
            os.startfile(r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance/year2_b.xlsx", )
    elif u_sem == "5" or u_sem == "6":
        if u_sec == "a":
            os.startfile(r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance/year3_a.xlsx", )
        else:
            os.startfile(r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance/year3_b.xlsx", )
    else:
        if u_sec == "a":
            os.startfile(r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance/year4_a.xlsx", )
        else:
            os.startfile(r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance/year4_b.xlsx", )


def attendance_info():
    sec = sec_view.get()
    sem = sem_view.get()
    s_sub = sub_view.get()
    # print(s_sub)
    s_usn = {}
    if sem == "1" or sem == "2":
        if sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year1_a.xlsx"
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year1_b.xlsx"

    elif sem == "3" or sem == "4":
        if sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year2_a.xlsx"
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year2_b.xlsx"
    elif sem == "5" or sem == "6":
        if sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year3_a.xlsx"
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year3_b.xlsx"
    else:
        if sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year4_a.xlsx"
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year4_b.xlsx"
    wb = load_workbook(year)
    # print(wb.sheetnames)
    sheet = wb.get_sheet_by_name(f"{s_sub}")
    r = 2
    names = {}
    while r < sheet.max_row:
        temp = sheet.cell(row=r, column=3).value
        c = 4
        s_usn[temp] = 0
        names[temp] = sheet.cell(row=r, column=2).value

        while c < (sheet.max_column + 1):
            if sheet.cell(row=r, column=c).value == "present":
                s_usn[temp] = s_usn[temp] + 1
            c = c + 1

        r = r + 1
    r = 2
    # print(s_usn)
    # print(names)
    lab1 = Label(view_label, bg="#FFF5EE")
    l1 = Listbox(lab1, width="100", height=20)
    l1.pack(pady=20)

    num = 3
    l1.insert(1,
              f"                                                                  Subject:{s_sub}                                                     Total class:{sheet.max_column - 2}                    ")
    l1.insert(2,
              "name                       usn                   class attended                                Percentage")
    for i in s_usn:
        per = (s_usn[i] / sheet.max_column) * 100
        l1.insert(num,
                  f"{names[i]}            {i}                   {s_usn[i]}                                                  {per}")
        num = num + 1
    view_frame1.pack_forget()
    lab1.pack()


def train_model():
    t.train()
    tmsg.showinfo("model training sucessful")


def message():
    sec = a_sec.get()
    sem = a_sem.get()
    s_sub = a_sub.get()
    if sem == "1" or sem == "2":
        if sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year1_a.xlsx"
            y = 'year1_a'
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year1_b.xlsx"
            y = 'year1_b'

    elif sem == "3" or sem == "4":
        if sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year2_a.xlsx"
            y = 'year2_a'
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year2_b.xlsx"
            y = 'year2_b'
    elif sem == "5" or sem == "6":
        if sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year3_a.xlsx"
            y = 'year3_a'
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year3_b.xlsx"
            y = 'year3_b'
    else:
        if sec == "a":
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year4_a.xlsx"
            y = 'year4_a'
        else:
            year = r"C:\Users\Suhil\PycharmProjects\tkinter\Attendance\year4_b.xlsx"
            y = 'year4_b'

    wb = load_workbook(year)
    sheet = wb.get_sheet_by_name(f"{s_sub}")
    wb1 = load_workbook("student.xlsx")
    sheet1 = wb1.get_sheet_by_name(f"{y}")
    r = 2
    name = []
    contact = []
    c = sheet.max_column
    while r < sheet.max_row:
        if sheet.cell(row=r, column=c).value == "Absent":
            name.append(sheet.cell(row=r, column=2).value)
            contact.append(sheet1.cell(row=r, column=6).value)
        r = r + 1
    exe1.send_message(name, s_sub, contact)
    tmsg.showinfo(message="message sent")


window = exe1.GUI()

# login page
l_login = Label(bg="silver")
f_login = Frame(l_login, pady="25", padx="25")
lb0 = Label(f_login, text="Enter Details", bg="orange", fg="blue", font="lucida 10 bold", width="35",
            pady="4").grid(
    columnspan=3, row=0, pady="15")
lb1 = Label(f_login, text="Enter ID: ", font="lucida 10 bold").grid(column=0, row=2, pady="4")
lid = StringVar()
e1 = Entry(f_login, textvariable=lid, width="28").grid(column=1, row=2)
lb2 = Label(f_login, text="Enter Password: ", font="lucida 10 bold").grid(column=0, row=3, pady="4")
l_pass = StringVar()
e2 = Entry(f_login, textvariable=l_pass, width="28").grid(column=1, row=3)
btn = Button(f_login, text="login", bg="green", fg="white", width="10", font="lucida 10 bold", command=verify)
btn.grid(columnspan=3, row=5, pady="10")
f_login.pack(pady=140)
l_login.pack()

# attendance page

attendance = Label(window)
attendance_label = Label(attendance)
attendance_frame1 = Frame(attendance_label, bg='silver', pady="25", padx="25")
b1 = Button(attendance_frame1, text="Take Attendance", fg="black", bg="#B6B6B4", width="15", font="lucida 10 bold",
            command=change1)
b1.pack(anchor="w", pady=15)
b2 = Button(attendance_frame1, text="View Files", fg="black", bg="silver", width="15", font="lucida 10 bold",
            command=change2)
b2.pack(anchor="w", pady=15)
attendance_frame1.pack(anchor="w", fill=BOTH, side="left")
attendance_label.pack(anchor="w", fill='both', side=LEFT, expand=FALSE)

attendance_label2 = Label(attendance, bg="#FEFCFF")
# write the code here
fd1 = Frame(attendance_label2, bg='silver', pady=25, padx=25)

lb1 = Label(fd1, text="Semester : ", bg='silver', font="lucida 10 bold").grid(column=0, row=2, pady="8")
a_sem = StringVar()
e1 = Entry(fd1, textvariable=a_sem, width="28").grid(column=1, row=2)
lb2 = Label(fd1, text="Section : ", bg='silver', font="lucida 10 bold").grid(column=0, row=3, pady="8")
a_sec = StringVar()
e2 = Entry(fd1, textvariable=a_sec, width="28").grid(column=1, row=3)
lb3 = Label(fd1, text="Subject : ", bg='silver', font="lucida 10 bold").grid(column=0, row=4, pady="8")
a_sub = StringVar()
e3 = Entry(fd1, textvariable=a_sub, width="28").grid(column=1, row=4)
btn = Button(fd1, text="Next", bg="green", fg="white", width="10", font="lucida 10 bold", command=change5)
btn.grid(columnspan=3, row=5, pady="10")
fd1.pack(pady=100)

fd = Frame(attendance_label2, pady="25", padx="25", bg='silver')
ld = Label(fd, text="This is Detect Section", bg="orange", fg="blue", font="lucida 10 bold", width="35", pady="4").grid(
    columnspan=3, row=0, pady="15")
b1 = Button(fd, text="Detect", bg="green", fg="white", width="15", font="lucida 10 bold", command=detect1)
b1.grid(columnspan=3, row=1, pady="10")
b2 = Button(fd, text="Update", bg="green", fg="white", width="15", font="lucida 10 bold", command=update1)
b2.grid(columnspan=3, row=3, pady="10")
b3 = Button(fd, text="Show", bg="green", fg="white", width="15", font="lucida 10 bold", command=show1)
b3.grid(columnspan=3, row=5, pady="10")
b5 = Button(fd, text="send message", bg="green", fg="white", width="15", font="lucida 10 bold", command=message)
b5.grid(columnspan=3, row=6, pady="10")
b4 = Button(fd, text="Back", bg="green", fg="white", width="15", font="lucida 10 bold", command=back)
b4.grid(columnspan=3, row=7, pady="10")

attendance_label2.pack(anchor="w", fill="both", side=LEFT, expand=TRUE)

view_label = Label(attendance, bg="silver")
view_frame1 = Frame(view_label, bg="#FFF5EE")
sem_view = StringVar()
sec_view = StringVar()
sub_view = StringVar()
student_sem = Label(view_frame1, text="Semester : ",bg="#FFF5EE", font="lucida 10 bold").grid(column=0, row=1,
                                                                                              pady="15")
sem = Entry(view_frame1, width="35", textvariable=sem_view)
sem.grid(column=1, row=1)
student_sec = Label(view_frame1, text="Section : ",bg="#FFF5EE", font="lucida 10 bold").grid(column=0, row=2,
                                                                                             pady="15")
sec = Entry(view_frame1, width="35", textvariable=sec_view)
sec.grid(column=1, row=2)
student_sub = Label(view_frame1, text="Subject : ",bg="#FFF5EE", font="lucida 10 bold").grid(column=0, row=3,
                                                                                             pady="15")
sub = Entry(view_frame1, width="35", textvariable=sub_view)
sub.grid(column=1, row=3)
get_info = Button(view_frame1, text="submit", bg="green", fg="white", width="10", font="lucida 10 bold",
                  command=attendance_info)
get_info.grid(columnspan=2, row=4, pady=10)
view_frame1.pack(pady=100)

attendance.pack(anchor="w", side="left", fill="both", expand=TRUE)

name_val = StringVar()
usn_val = StringVar()
sec_val = StringVar()
sem_val = StringVar()
phone_val = StringVar()
email_val = StringVar()

student = Label(window)

student_frame1 = Frame(student, bg="silver", pady="25", padx="25")
b1 = Button(student_frame1, text="Register", fg="black", bg="silver", width="15", font="lucida 10 bold",
            command=change3)
b1.pack(anchor="w", pady=15)
b2 = Button(student_frame1, text="Students Info", fg="black", bg="silver", width="15", font="lucida 10 bold",
            command=get_student_info)

b2.pack(anchor="w", pady=15)
b3 = Button(student_frame1, text="Train Model", fg="black", bg="silver", width="15", font="lucida 10 bold",
            command=train_model)
b3.pack(anchor="w", pady=15)
student_frame1.pack(anchor="w", fill=BOTH, side="left")

student_label1 = Label(student, bg="white")
f1 = Frame(student_label1, bg="silver", padx="25")
l0 = Label(f1, text="Registration Form", bg="orange", fg="blue", font="lucida 15 bold", width="35", pady="4").grid(
    columnspan=3, row=0, pady="15")
l1 = Label(f1, text="Name : ", bg="silver", font="lucida 10 bold").grid(column=0, row=1, pady="4")
name = Entry(f1, width="35", textvariable=name_val)
name.grid(column=1, row=1)
l2 = Label(f1, text="USN : ", bg="silver", font="lucida 10 bold").grid(column=0, row=2, pady="4")
usn = Entry(f1, width="35", textvariable=usn_val)
usn.grid(column=1, row=2)

l32 = Label(f1, text="Section : ", bg="silver", font="lucida 10 bold").grid(column=0, row=4, pady="4")
section = Entry(f1, width="35", textvariable=sec_val)
section.grid(column=1, row=4)

l33 = Label(f1, text="Sem : ", bg="silver", font="lucida 10 bold").grid(column=0, row=5, pady="4")
semester = Entry(f1, width="35", textvariable=sem_val)
semester.grid(column=1, row=5)

l5 = Label(f1, text="Contact No : ", bg="silver", font="lucida 10 bold").grid(column=0, row=6, pady="4")
contact = Entry(f1, width="35", textvariable=phone_val)
contact.grid(column=1, row=6)

l6 = Label(f1, text="Email : ", bg="silver", font="lucida 10 bold").grid(column=0, row=7, pady="6")
email = Entry(f1, width="35", textvariable=email_val)
email.grid(column=1, row=7)
btn = Button(f1, text="Take photo", bg="green", fg="white", width="10", font="lucida 10 bold", command=take_photo)
btn.grid(column=1, row=8, pady="10")
btn1 = Button(f1, text="Submit", bg="green", fg="white", width="10", font="lucida 10 bold", command=write)
btn1.grid(column=2, row=8, pady="10")
f1.pack(pady="90", padx=20)
student_label1.pack(anchor="w", fill='both', side=LEFT, expand=TRUE)

student_label2 = Label(student, bg="white")

window.mainloop()
