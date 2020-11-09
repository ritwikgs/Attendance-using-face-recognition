import os
import cv2
from time import sleep
import datetime
from openpyxl import load_workbook
import openpyxl


def detect(sem, sec):
    fname = 'recognizer/trainingData.yml'
    if not os.path.isfile(fname):
        print('first train the data')
        exit(0)

    names = {}
    labels = []
    students = []
    if sem == "1" or sem == "2":
        if sec == "a":
            year = "year1_a"
        else:
            year = "year1_b"

    elif sem == "3" or sem == "4":
        if sec == "a":
            year = "year2_a"
        else:
            year = "year2_b"
    elif sem == "5" or sem == "6":
        if sec == "a":
            year = "year3_a"
        else:
            year = "year3_b"
    else:
        if sec == "a":
            year = "year4_a"
        else:
            year = "year4_b"

    wb = load_workbook(f"Attendance\{year}.xlsx")
    sheets = wb.active
    row = 2
    while row < sheets.max_row:
        n = sheets.cell(row, 2).value
        usn = sheets.cell(row, 3).value
        row = row + 1
        temp = usn
        usn = temp[3:5]
        usn = usn + temp[-3:]
        # print(n)
        # print(usn)
        names[int(usn)] = n
    # print(names)
    face_recognizer = cv2.face.LBPHFaceRecognizer_create()
    face_cascade = cv2.CascadeClassifier(r'C:\Users\Suhil\PycharmProjects\tkinter\HaarCascade\haarcascade_frontalface_default.xml')
    face_recognizer.read(fname)

    num = 0
    justlabels = set()

    while num <= 10:
        cap = cv2.VideoCapture(0)

        ret, img = cap.read()
        # print(ret)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 3)
            label, confidence = face_recognizer.predict(gray[y:y + h, x:x + w])
            # print('label:', label)
            # print('confidence:', confidence)
            predicted_name = names[label]
            if confidence < 90:
                confidence = 100 - round(confidence)
                cv2.putText(img, predicted_name + str(confidence) + '%', (x + 2, y + h - 4), cv2.FONT_HERSHEY_SIMPLEX,
                            1, (150, 255, 0), 2)
                labels.append(label)
                students.append(names[label])
                totalstudents = set(students)
                justlabels = set(labels)
                # print('student Recognised : ', totalstudents, justlabels)

            cv2.imshow('Face Recognizer', img)
            cv2.waitKey(100)

        num = num + 1
        # if num > 10:
        #     cap.release()
        #     sleep(4)
        #     print('we are done!')
        #     break
        sleep(5)
    cap.release()

    cv2.destroyAllWindows()
    return justlabels
